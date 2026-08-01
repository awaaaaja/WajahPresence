"""Export laporan absensi (Sprint 4, Prompt 4.5) — Excel & PDF.

Murni helper: menerima baris log (list[dict]) dan menghasilkan bytes file.
Format kolom konsisten antar format supaya mudah dibaca non-teknis.
"""

from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any, Iterator

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

_TITLE_STYLE = ParagraphStyle(name="title", fontSize=10, leading=13, spaceAfter=6)

EXPORT_COLUMNS = [
    ("timestamp", "Waktu"),
    ("nama", "Nama"),
    ("email", "Email"),
    ("status", "Status"),
    ("confidence_score", "Confidence"),
    ("lat", "Latitude"),
    ("lng", "Longitude"),
    ("site", "Site"),
    ("gps_accuracy", "GPS Accuracy (m)"),
    ("ip_address", "IP Address"),
    ("ip_mismatch_flag", "IP Mismatch"),
    ("rejection_reason", "Alasan"),
    ("reviewed_at", "Direview"),
]

STATUS_LABEL = {"success": "Sukses", "rejected": "Ditolak", "suspicious": "Mencurigakan"}


def _fmt_value(key: str, value: Any) -> Any:
    if value is None:
        return ""
    if key == "timestamp" or key == "reviewed_at":
        try:
            return datetime.fromisoformat(str(value).replace("Z", "+00:00")).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            return str(value)
    if key == "status":
        return STATUS_LABEL.get(str(value), str(value))
    if key == "ip_mismatch_flag":
        return "Ya" if value else "Tidak"
    if isinstance(value, float):
        return round(value, 6)
    return value


def _rows(rows: list[dict[str, Any]]) -> list[list[Any]]:
    return [[_fmt_value(k, r.get(k)) for k, _ in EXPORT_COLUMNS] for r in rows]


def build_xlsx(rows: list[dict[str, Any]], title: str) -> bytes:
    # write_only jauh lebih cepat utk ribuan baris (tanpa cell-object per sel)
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Laporan Absensi")

    header = [label for _, label in EXPORT_COLUMNS]
    ws.append([_styled_header_cell(ws, h) for h in header])

    for row in _rows(rows):
        ws.append(row)

    widths = {key: 30 if key in ("nama", "email", "rejection_reason") else 14 for key, _ in EXPORT_COLUMNS}
    for i, (key, _) in enumerate(EXPORT_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths[key]

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _styled_header_cell(ws: Any, value: str) -> WriteOnlyCell:
    cell = WriteOnlyCell(ws, value=value)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")
    return cell


def build_pdf(rows: list[dict[str, Any]], title: str) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=10 * mm,
        leftMargin=10 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=title,
    )

    header = [label for _, label in EXPORT_COLUMNS]
    body = ([str(v) if v is not None else "" for v in row] for row in _rows(rows))

    # Tabel raksasa (> ribuan baris) membuat reportlab lambat —
    # pecah per 300 baris supaya layout tiap halaman ringan.
    chunk_size = 300
    tables = [
        Table(
            [header, *chunk],
            colWidths=[
                34 * mm, 32 * mm, 36 * mm, 20 * mm, 18 * mm,
                20 * mm, 20 * mm, 26 * mm, 24 * mm, 24 * mm,
                18 * mm, 45 * mm, 34 * mm,
            ],
            repeatRows=1,
        )
        for chunk in _chunks(body, chunk_size)
    ]

    generated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    story = [
        Paragraph(f"<b>{title}</b> — dibuat {generated}", _TITLE_STYLE),
        Spacer(1, 4 * mm),
    ]
    for table in tables:
        table.setStyle(
            TableStyle(
                [
                    ("FONTSIZE", (0, 0), (-1, -1), 6.5),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#d1d5db")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f3f4f6")]),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.append(table)
        story.append(Spacer(1, 4 * mm))
    doc.build(story)
    return buffer.getvalue()


def _chunks(iterable: Any, size: int) -> Iterator[list[Any]]:
    chunk: list[Any] = []
    for item in iterable:
        chunk.append(item)
        if len(chunk) == size:
            yield chunk
            chunk = []
    if chunk:
        yield chunk
